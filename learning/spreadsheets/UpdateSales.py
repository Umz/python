#!  python3
#   UpdateSales.py - Update the sales spreadsheet correcting values

import openpyxl

#   LOAD the workbook and the sheet with the data

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

#   DICTIONARY of products with updated prices

PRICE_UPDATES = {
    'Garlic':3.07,
    'Celery':1.19,
    'Lemon':1.27
}

#   LOOP through the rows to change prices of dictionary items

for row in range(2, sheet.max_row):
    product = sheet.cell(row=row, column=1).value
    if product in PRICE_UPDATES:
        sheet.cell(row=row, column=2).value = PRICE_UPDATES[product]

sheet.row_dimensions[1].height = 40     # SET row height for header
sheet.freeze_panes = 'A2'               # FREEZE first row

wb.save('produceSales_new.xlsx')
